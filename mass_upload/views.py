import datetime
import io
import os
from django.http import FileResponse, HttpResponse
from django.shortcuts import redirect, render
from django.contrib import messages,auth
import openpyxl
from  django.contrib.auth.decorators import login_required,user_passes_test
# import slugify
from django.template.defaultfilters import slugify
from django.db.models.functions import Now
from pytz import utc
from accounts.views import check_role_vendor
from store.models import Brand
from category.models import Category, Category_Upload,  Fourth_Level_Category, Main_Category, Sub_Category, Vendor_Category, Vendor_Main_Category, Vendor_Sub_Category
from mass_upload.forms import UploadFileForm
from mass_upload.models import ProductRecord, StgProduct
import datetime
# Create your views here.
from django.db.models import Q
from tablib import Dataset

from store.models import Color, Images, Product, Size, Variants

# from openpyxl import Workbook

from vendor.models import Vendor
from django.shortcuts import render
from openpyxl import load_workbook
from .models import Product, Stock_keeping_unit

# def ImportExcel2(request):
#      if request.method == 'POST' :
#           product_resource = ProductResource()
#           dataset = Dataset()
#           new_products = request.Files['my_file']

#           if not new_products.name.endswith('xlsx'):
#                messages.info(request,'Wrong format')
#                return render(request,'store/bulkproduct.html')
#           imported_data = dataset.load(new_products.read(),format='xlsx')
#           for data in imported_data:
#                value = Product(
#                     data[0],
#                     data[1],
#                     data[2]
#                )
#                value.save()
#      return render(request,'store/bulkproduct.html')

def getcolor(colorname):
     return Color.objects.get(name=colorname)

def getsize(sizename):
     return Size.objects.get(name=sizename)

def getmaincategory(maincat_id):
     return Main_Category.objects.get(id=maincat_id)

@login_required(login_url='login')
@user_passes_test(check_role_vendor)
def bulk_product(request):
    uploadform=UploadFileForm(None)
    vendor = Vendor.objects.get(user=request.user)
    if request.method == 'POST' :
        uploadform = UploadFileForm(request.POST, request.FILES)
        if uploadform.is_valid():
            # file = uploadform.cleaned_data['file']
            file = request.FILES['file']
            workbook = openpyxl.load_workbook(filename=file, read_only=True)
            
            # Get name of the first sheet and then open sheet by name
            first_sheet = workbook.get_sheet_names()[0]
            worksheet = workbook.get_sheet_by_name(first_sheet)

            data = []
            print('1')
            try:
                print('2')
                if StgProduct.objects.filter(file_name=file).exists():
                     messages.info(request,'File with same name is already uploaded')
                     return redirect('bulk_product')
                else:
                    for row in worksheet.iter_rows(): # Offset for header
                            stgproduct = StgProduct()
                        
                        
                            stgproduct.vendor = vendor
                            stgproduct.category_id =row[0].value
                            stgproduct.product_name =row[1].value
                            stgproduct.featured_image=row[2].value
                            stgproduct.image1 =row[3].value
                            stgproduct.image2=row[4].value
                            stgproduct.image3=row[5].value
                            stgproduct.image4=row[6].value
                            stgproduct.image5=row[7].value
                            stgproduct.image6=row[8].value
                            stgproduct.image7=row[9].value
                            stgproduct.image8=row[10].value
                            stgproduct.brand =row[11].value
                            stgproduct.product_information =row[12].value
                            stgproduct.variant =row[13].value
                            stgproduct.variant_title=row[14].value
                            stgproduct.variant_color =row[15].value
                            stgproduct.variant_size =row[16].value
                            stgproduct.image_variant =row[17].value
                            stgproduct.variant_stock =row[18].value
                            stgproduct.variant_price =row[19].value
                            stgproduct.variant_discount =row[20].value
                            stgproduct.variant_max_allowed_quantity =row[21].value
                            stgproduct.stock =row[22].value
                            stgproduct.price =row[23].value
                            stgproduct.discount =row[24].value
                            stgproduct.weight =row[25].value
                            stgproduct.parcel_size_L =row[26].value
                            stgproduct.parcel_size_W =row[27].value
                            stgproduct.parcel_size_H =row[28].value
                            stgproduct.max_allowed_quantity =row[29].value
                            stgproduct.min_order_quantity =row[30].value
                            stgproduct.file_name = file

                            data.append(stgproduct)
                    print('3')
                    StgProduct.objects.bulk_create(data)
                    print('4')
            except Exception as e:
                print(f'exception {e}')
                messages.error(request,"Error",extra_tags="excelerror")

            upload_cnt = 0
            # Product.objects.filter(id__gte=19).delete()
            # Images.objects.filter(product_id__gte=19).delete()
            # Variants.objects.filter(product_id__gte=19).delete()
            # Variants.objects.filter(product_id__gte=19).delete()
            try:
                StgProduct.objects.filter(file_name=file,category_id__in=['Category','Mandatory']).delete()
                stgdata = StgProduct.objects.filter(file_name=file)
                records = stgdata.count()
                for row in stgdata :
                        variant = ''
                        sku = ''
                        stgprod = StgProduct.objects.get(id=row.id)
                        try :
                            if StgProduct.objects.filter(Q(variant__isnull=True),id=row.id).exists() :
                                stgprod.status='Error'
                                stgprod.processed=False
                                stgprod.error = 'Variant cannot be blank'
                                stgprod.save()
                            elif  StgProduct.objects.filter(Q(category_id__isnull=True)| Q(product_name__isnull=True)| Q(featured_image__isnull=True)| Q(brand__isnull=True)| Q(variant_title__isnull=True)| Q(variant_color__isnull=True)| Q(variant_size__isnull=True)| Q(image_variant__isnull=True)| Q(variant_stock__isnull=True)| Q(variant_price__isnull=True)| Q(stock__isnull=True)| Q(price__isnull=True)| Q(weight__isnull=True)| Q(max_allowed_quantity__isnull=True),id=row.id, variant='Size-Color').exists() :
                                stgprod.status='Error'
                                stgprod.processed=False
                                stgprod.error = 'Please enter mandatoey fields : category, product name, featured image, brand,	variation, variation name, variation color, variation size, variation image, variation stock,variation price, stock, price, package weight, max allowed quantity'
                                stgprod.save()
                            elif  StgProduct.objects.filter(Q(category_id__isnull=True)| Q(product_name__isnull=True)| Q(featured_image__isnull=True)| Q(brand__isnull=True)| Q(variant_title__isnull=True)| Q(variant_color__isnull=True)| Q(image_variant__isnull=True)| Q(variant_stock__isnull=True)| Q(variant_price__isnull=True)| Q(stock__isnull=True)| Q(price__isnull=True)| Q(weight__isnull=True)| Q(max_allowed_quantity__isnull=True),id=row.id, variant='Color').exists() :
                                stgprod.status='Error'
                                stgprod.processed=False
                                stgprod.error = 'Please enter mandatoey fields : category, product name, featured image, brand,	variation, variation name, variation color, variation image, variation stock,variation price, stock, price, package weight, max allowed quantity'
                                stgprod.save()
                            elif  StgProduct.objects.filter(Q(category_id__isnull=True)| Q(product_name__isnull=True)| Q(featured_image__isnull=True)| Q(brand__isnull=True)| Q(variant_title__isnull=True)| Q(variant_size__isnull=True)| Q(image_variant__isnull=True)| Q(variant_stock__isnull=True)| Q(variant_price__isnull=True)| Q(stock__isnull=True)| Q(price__isnull=True)| Q(weight__isnull=True)| Q(max_allowed_quantity__isnull=True),id=row.id, variant='Size').exists() :
                                stgprod.status='Error'
                                stgprod.processed=False
                                stgprod.error = 'Please enter mandatoey fields : category, product name, featured image, brand,	variation, variation name,  variation size, variation image, variation stock,variation price, stock, price, package weight, max allowed quantity'
                                stgprod.save()
                            elif StgProduct.objects.filter(Q(category_id__isnull=True)| Q(product_name__isnull=True)| Q(featured_image__isnull=True)| Q(brand__isnull=True)| Q(stock__isnull=True)| Q(price__isnull=True)| Q(weight__isnull=True)| Q(max_allowed_quantity__isnull=True),id=row.id, variant='None').exists() :
                                stgprod.status='Error'
                                stgprod.processed=False
                                stgprod.error = 'Please enter mandatory fields : category, product name, featured image, brand,	variation,stock, price, package weight, max allowed quantity'
                                stgprod.save()
                            else:
                                    try:
                                                print('5')
                                                sub_category = ''
                                                category = ''
                                                fourth_level_category = ''
                                                category_upload = Category_Upload.objects.get(id = row.category_id)
                                                print('6')
                                                if category_upload.main_category_id :
                                                    main_category = Main_Category.objects.get(id=category_upload.main_category_id)
                                                print('7')  
                                                if category_upload.category_id :
                                                    category = Category.objects.get(id=category_upload.category_id,main_category=main_category)
                                                print('8')
                                                if category_upload.sub_category_id :
                                                    sub_category = Sub_Category.objects.get(id=category_upload.sub_category_id,category=category)
                                                print('9')
                                                if category_upload.fourth_level_category_id :
                                                    fourth_level_category = Fourth_Level_Category.objects.get(id=category_upload.fourth_level_category_id,sub_category=sub_category)
                                                print('10')

                                                
                                                
                                    except Category_Upload.DoesNotExist:
                                                pass
                                    
                                    try:
                                            if Product.objects.filter(product_name=row.product_name,vendor=vendor).exists():
                                                    product = Product.objects.get(product_name=row.product_name,vendor=vendor)
                                                    stgprod.status='Success'
                                                    stgprod.processed=True
                                                    stgprod.error = 'Product already exists '
                                                    stgprod.save()
                                            else :   
                                                        #create product
                                                        print('11')
                                                        slug = slugify(row.product_name)
                                                        print('12')
                                                        product = Product.objects.create(
                                                            vendor = vendor,
                                                            slug = slug ,
                                                            main_category = main_category,
                                                            product_name = row.product_name,
                                                            featured_image = row.featured_image,
                                                            variant = row.variant,
                                                            stock = row.stock,  
                                                            price = row.price,
                                                            weight = row.weight,
                                                            max_allowed_quantity = row.max_allowed_quantity
                                                        )
                                                        print('13')
                                                        # product = Product.objects.get(product_name=row.product_name)
                                                        print('14') 

                                                        if category :
                                                            product.category=category
                                                        if sub_category :
                                                            product.sub_category = sub_category
                                                        if  fourth_level_category:
                                                            product.fourth_level_category=fourth_level_category
                                                        if row.brand:
                                                            product.brand = Brand.objects.get(name=row.brand)

                                                        if row.parcel_size_L:
                                                            product.parcel_size_L = row.parcel_size_L
                                                        if row.parcel_size_W:
                                                            product.parcel_size_W = row.parcel_size_W
                                                        if row.parcel_size_H:
                                                            product.parcel_size_H = row.parcel_size_H
                                                        if row.product_information:
                                                            product.product_information = row.product_information
                                                        if row.discount:
                                                            product.discount = row.discount
                                                        if row.min_order_quantity:
                                                            product.min_order_quantity = row.min_order_quantity

                                                        product.save()
                                                        print('15')

                                                        #add vendor categories

                                                        try:
                                                            obj = Vendor_Main_Category.objects.get(vendor=product.vendor,name= product.main_category.name)
                                                        except Vendor_Main_Category.DoesNotExist:
                                                            Vendor_Main_Category.objects.create(name=product.main_category.name,
                                                                                    vendor=product.vendor,
                                                                                    slug = slugify(product.main_category.name))
                                                        vend_maincat = Vendor_Main_Category.objects.get(vendor=product.vendor,name= product.main_category.name)
                                                        try:
                                                            obj_cat = Vendor_Category.objects.get(vendor=product.vendor,name=product.category.name,main_category_id=vend_maincat.id)
                                                        except Vendor_Category.DoesNotExist:
                                                            Vendor_Category.objects.create(name=product.category.name,
                                                                                        main_category_id = vend_maincat.id,
                                                                                    vendor=product.vendor,
                                                                                    slug = slugify(product.category.name))
                                                        cat = Vendor_Category.objects.get(vendor=product.vendor,name=product.category.name,main_category_id=vend_maincat.id)
                                                        try:
                                                            obj_subcat = Vendor_Sub_Category.objects.get(vendor=product.vendor,name=product.sub_category.name,category_id=product.category.id,main_category_id=vend_maincat.id)
                                                        except Vendor_Sub_Category.DoesNotExist:
                                                            Vendor_Sub_Category.objects.create(name=product.sub_category.name,
                                                                                    vendor=product.vendor,
                                                                                    main_category_id = vend_maincat.id,
                                                                                    category_id = cat.id,
                                                                                    slug = slugify(product.sub_category.name))
                                                        #add product images
                                                        if row.image1 :
                                                            Images.objects.create(product=product, image=row.image1)
                                                        print('16')
                                                        if row.image2 :
                                                            Images.objects.create(product=product, image=row.image2)
                                                        print('17')
                                                        if row.image3 :
                                                            Images.objects.create(product=product, image=row.image3)
                                                        print('18')

                                                        if row.image4 :
                                                            Images.objects.create(product=product, image=row.image4)
                                                        print('19')

                                                        if row.image5 :
                                                            Images.objects.create(product=product, image=row.image5)
                                                        print('20')

                                                        if row.image6 :
                                                            Images.objects.create(product=product, image=row.image6)
                                                        print('21')

                                                        if row.image7 :
                                                            Images.objects.create(product=product, image=row.image7)
                                                        print('22')

                                                        if row.image8 :
                                                            Images.objects.create(product=product, image=row.image8)
                                                        print('23')

                                            #add variants
                                            if product.variant != 'None' :
                                                            
                                                            if product.variant == 'Color':
                                                                    color = getcolor(row.variant_color)
                                                                    if Variants.objects.filter(product=product,color=color).exists():
                                                                        stgprod.status='Failed'
                                                                        stgprod.processed=False
                                                                        stgprod.error =  'Product with same color variant already available '
                                                                        stgprod.save()
                                                                        print('24')
                                                                    else:
                                                                        variant = Variants.objects.create(product = product,
                                                                                title = row.variant_title,
                                                                                color = getcolor(row.variant_color),
                                                                                variant_stock = row.variant_stock,
                                                                                variant_price = row.variant_price,
                                                                                image_variant = row.image_variant)
                                                                        print('25')
                                                                        if row.variant_discount:
                                                                            variant.variant_discount = row.variant_discount
                                                                            print('27')
                                                                        if row.variant_max_allowed_quantity :
                                                                            variant.variant_max_allowed_quantity = row.variant_max_allowed_quantity
                                                                            print('28')
                                                                        variant.save()
                                                                        print('29')
                                                                        Stock_keeping_unit.objects.create(product =  product,
                                                                                            vendor = vendor,
                                                                                            variant_id = variant.id)
                                                                        print('30')
                                                                        if product.is_active == False :
                                                                            product.is_active = True
                                                                            product.save()
                                                                        print('31')
                                                            elif product.variant == 'Size':
                                                                    size = getsize(row.variant_size)
                                                                    
                                                                    if Variants.objects.filter(product=product,size=size).exists() :
                                                                        stgprod.status='Failed'
                                                                        stgprod.processed=False
                                                                        stgprod.error =  'Product with same size variant already available '
                                                                        stgprod.save()
                                                                        print('32')
                                                                    else:
                                                                        variant = Variants.objects.create(product = product,
                                                                                title = row.variant_title,
                                                                                size = getsize(row.variant_size),
                                                                                variant_stock = row.variant_stock,
                                                                                variant_price = row.variant_price,
                                                                                image_variant = row.image_variant)
                                                                        print('33')
                                                                        if row.variant_discount:
                                                                            variant.variant_discount = row.variant_discount
                                                                        if row.variant_max_allowed_quantity :
                                                                            variant.variant_max_allowed_quantity = row.variant_max_allowed_quantity
                                                                        variant.save()
                                                                        Stock_keeping_unit.objects.create(product =  product,
                                                                                            vendor = vendor,
                                                                                            variant = variant)
                                                                        if product.is_active == False :
                                                                            product.is_active = True
                                                                            product.save()
                                                                        print('34')
                                                            elif product.variant == 'Size-Color':
                                                                    size = getsize(row.variant_size)
                                                                    color = getcolor(row.variant_color)
                                                                    if Variants.objects.filter(product=product,size=size,color = color).exists() :
                                                                            stgprod.status='Failed'
                                                                            stgprod.processed=False
                                                                            stgprod.error =  'Product with same size and color variant already available '
                                                                            stgprod.save()
                                                                            print('35')
                                                                    else:
                                                                        variant = Variants.objects.create(product = product,
                                                                                title = row.variant_title,
                                                                                size = getsize(row.variant_size),
                                                                                color = getcolor(row.variant_color),
                                                                                variant_stock = row.variant_stock,
                                                                                variant_price = row.variant_price,
                                                                                image_variant = row.image_variant)
                                                                        print('36')
                                                                        if row.variant_discount:
                                                                            variant.variant_discount = row.variant_discount
                                                                        if row.variant_max_allowed_quantity :
                                                                            variant.variant_max_allowed_quantity = row.variant_max_allowed_quantity
                                                                        variant.save()
                                                                        Stock_keeping_unit.objects.create(product =  product,
                                                                                            vendor = vendor,
                                                                                            variant = variant)
                                                                        if product.is_active == False :
                                                                            product.is_active = True
                                                                            product.save()
                                                                        print('37')
                                            else: 
                                                            product.is_active = True
                                                            product.save()
                                                            try :
                                                                sku =  Stock_keeping_unit.objects.get(product=product,vendor=vendor)
                                                                print('38')
                                                            except :      
                                                                Stock_keeping_unit.objects.create(product =  product,
                                                                                            vendor = vendor)
                                            
                                                                print('39')

                                            
                                                    
                                            print('successfully processed')
                                            stgprod.status =  'Success'
                                            stgprod.processed=True
                                            stgprod.save()
                                            print('40')

                                            
                                    except Exception as e:
                                            if product :
                                                    product.delete()
                                            if Images.objects.filter(product=product).exists():
                                                        image = Images.objects.filter(product=product)
                                                        image.delete()
                                            if Variants.objects.filter(product=product).exists():
                                                        variants = Variants.objects.filter(product=product)
                                                        variants.delete()
                                            if sku :
                                                    sku.delete()
                                    
                                            print(f'exception {e}')
                                            stgprod.status='Failed'
                                            stgprod.processed=False
                                            stgprod.error=e
                                            stgprod.save()
                                            print('41')
                                             
                                         
                        except Exception as e:
                                stgprod.status='Failed'
                                stgprod.processed=False
                                stgprod.error=e
                                stgprod.save()
                             
                                                
                stgprod_cnt =  StgProduct.objects.filter(file_name=file,processed=True,status='Success').count()
                ProductRecord.objects.create(vendor=vendor,
                                         file_name = file,
                                         processed = stgprod_cnt,
                                         products = records)
                pr = ProductRecord.objects.get(file_name = file,vendor=vendor)
                if pr.processed > 0 & pr.processed < pr.products :
                     pr.status = 'Partially Successful'
                     pr.save()
                elif  pr.processed == 0 :
                     pr.status = 'Unsuccessful'
                     pr.save()
                elif pr.processed == pr.products :
                     pr.status = 'Successful'
                     pr.save()
                messages.info(request, 'File is uploaded successfully')
            
            except Exception as e:
                print(f'exception {e}')
                stgprod_cnt =  StgProduct.objects.filter(file_name=file,processed=True,status='Success').count()
                ProductRecord.objects.create(vendor=vendor,
                                         file_name = file,
                                         processed = stgprod_cnt,
                                         products = records)
                pr = ProductRecord.objects.get(file_name = file,vendor=vendor)
                if pr.processed > 0 & pr.processed < pr.products :
                     pr.status = 'Partially Successful'
                     pr.save()
                elif  pr.processed == 0 :
                     pr.status = 'Unsuccessful'
                     pr.save()
                elif pr.processed == pr.products :
                     pr.status = 'Successful'
                     pr.save()
            
                messages.error(request,"Error",extra_tags="excelerror")   

            report_data = ProductRecord.objects.filter(vendor=vendor)
            context = {
                 'report_data' : report_data,
                 'uploadform' : uploadform,
                }
    else:    
        report_data = ProductRecord.objects.filter(vendor=vendor)
        context = {
            'report_data' : report_data,
            'uploadform' : uploadform,
        }
    
    return render(request, 'mass_upload/bulk_product.html',context)

import xlsxwriter

@login_required(login_url='login')
@user_passes_test(check_role_vendor)
def download_basic_template(request):
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define file name
    filename = 'products.xlsx'
    # Define the full file path
    filepath = BASE_DIR + '/mass_upload/' + filename
    return FileResponse(open(filepath, 'rb'), as_attachment=True)

from datetime import datetime, timezone
def utc_to_local(utc_dt):
    return utc_dt.replace(tzinfo=timezone.utc).astimezone(tz=None)

@login_required(login_url='login')
@user_passes_test(check_role_vendor)
def generate(request,file_name):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'remove_timezone': True})
    worksheet = workbook.add_worksheet('test')

    # cell_format = workbook.add_format({'bold': True, 'italic': True})


    # worksheet.write(0, 0, 'name',cell_format)
    # worksheet.write(0, 1, 'category',cell_format)

    #daya start

    bold = workbook.add_format({'bold': True})

    bold.set_pattern(1)  # This is optional when using a solid fill.
    bold.set_bg_color('orange')


    # Some data we want to write to the worksheet.
    reporte = StgProduct.objects.filter(file_name = file_name) #my model

    # Start from the first cell. Rows and columns are zero indexed.
    row = 1
    col = 0


    # Iterate over the data and write it out row by row.
    for linea in reporte:

        # naive_datetime = utc_to_local(linea.created_date)
        # naive_datetime = utc_datetime.replace(tzinfo=None)


        worksheet.write(row, col,linea.category_id)
        worksheet.write(row, col +  1, linea.product_name)
        worksheet.write(row, col + 2, linea.featured_image)
        worksheet.write(row, col + 3, linea.image1)
        worksheet.write(row, col + 4, linea.image2)
        worksheet.write(row, col + 5, linea.image3)
        worksheet.write(row, col + 6, linea.image4)
        worksheet.write(row, col + 7, linea.image5)
        worksheet.write(row, col + 8, linea.image6)
        worksheet.write(row, col + 9, linea.image7)
        worksheet.write(row, col + 10, linea.image8)
        worksheet.write(row, col + 11, linea.brand)
        worksheet.write(row, col + 12, linea.product_information)
        worksheet.write(row, col + 13, linea.variant)
        worksheet.write(row, col + 14, linea.variant_title)
        worksheet.write(row, col + 15, linea.variant_color)
        worksheet.write(row, col + 16, linea.variant_size)
        worksheet.write(row, col + 17, linea.image_variant)
        worksheet.write(row, col + 18, linea.variant_stock)
        worksheet.write(row, col + 19, linea.variant_price)
        worksheet.write(row, col + 20, linea.variant_discount)
        worksheet.write(row, col + 21, linea.variant_max_allowed_quantity)
        worksheet.write(row, col + 22, linea.stock)
        worksheet.write(row, col + 23, linea.price)
        worksheet.write(row, col + 24, linea.discount)
        worksheet.write(row, col + 25, linea.weight)
        worksheet.write(row, col + 26, linea.parcel_size_L)
        worksheet.write(row, col + 27, linea.parcel_size_W)
        worksheet.write(row, col + 28, linea.parcel_size_H)
        worksheet.write(row, col + 29, linea.max_allowed_quantity)
        worksheet.write(row, col + 30, linea.min_order_quantity)
        worksheet.write(row, col + 31, linea.file_name)
        worksheet.write(row, col + 32, linea.created_date)
        worksheet.write(row, col + 33, linea.processed)
        worksheet.write(row, col + 34, linea.status)
        worksheet.write(row, col + 35, linea.error)
        row += 1
    
    worksheet.write(0,0,'category_id', bold)
    worksheet.write(0,1,'product_name', bold)
    worksheet.write(0,2,'featured_image', bold)
    worksheet.write(0,3,'image1', bold)
    worksheet.write(0,4,'image2', bold)
    worksheet.write(0,5,'image3', bold)
    worksheet.write(0,6,'image4', bold)
    worksheet.write(0,7,'image5', bold)
    worksheet.write(0,8,'image6', bold)
    worksheet.write(0,9,'image7', bold)
    worksheet.write(0,10,'image8', bold)
    worksheet.write(0,11,'brand', bold)
    worksheet.write(0,12,'product_information', bold)
    worksheet.write(0,13,'variant', bold)
    worksheet.write(0,14,'variant_title', bold)
    worksheet.write(0,15,'variant_color', bold)
    worksheet.write(0,16,'variant_size', bold)
    worksheet.write(0,17,'image_variant', bold)
    worksheet.write(0,18,'variant_stock', bold)
    worksheet.write(0,19,'variant_price', bold)
    worksheet.write(0,20,'variant_discount', bold)
    worksheet.write(0,21,'variant_max_allowed_quantity', bold)
    worksheet.write(0,22,'stock', bold)
    worksheet.write(0,23,'price', bold)
    worksheet.write(0,24,'discount', bold)
    worksheet.write(0,25,'weight', bold)
    worksheet.write(0,26,'parcel_size_L', bold)
    worksheet.write(0,27,'parcel_size_W', bold)
    worksheet.write(0,28,'parcel_size_H', bold)
    worksheet.write(0,29,'max_allowed_quantity', bold)
    worksheet.write(0,30,'min_order_quantity', bold)
    worksheet.write(0,31,'file_name', bold)
    worksheet.write(0,32,'created_date', bold)
    worksheet.write(0,33,'processed', bold)
    worksheet.write(0,34,'status', bold)
    worksheet.write(0,35,'error', bold)

    #daya end

    workbook.close()
    output.seek(0)

    filename = 'django_simple.xlsx'
    response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    response['Content-Disposition'] = 'attachment; filename=%s' % filename

    return response



@login_required(login_url='login')
@user_passes_test(check_role_vendor)
def generate_subcat(request):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'remove_timezone': True})
    worksheet = workbook.add_worksheet('test')

    # cell_format = workbook.add_format({'bold': True, 'italic': True})


    # worksheet.write(0, 0, 'name',cell_format)
    # worksheet.write(0, 1, 'category',cell_format)

    #daya start

    bold = workbook.add_format({'bold': True})

    bold.set_pattern(1)  # This is optional when using a solid fill.
    bold.set_bg_color('orange')


    # Some data we want to write to the worksheet.
    reporte = Sub_Category.objects.all().order_by('id') #my model

    # Start from the first cell. Rows and columns are zero indexed.
    row = 1
    col = 0


    # Iterate over the data and write it out row by row.
    for linea in reporte:

        # naive_datetime = utc_to_local(linea.created_date)
        # naive_datetime = utc_datetime.replace(tzinfo=None)

        worksheet.write(row, col, linea.id)
        worksheet.write(row, col +  1, linea.name)
        worksheet.write(row, col +  2, linea.slug)

        worksheet.write(row, col +  3 ,linea.category_id)
        row += 1
    
    worksheet.write(0,0,'id', bold)
    worksheet.write(0,1,'name', bold)
    worksheet.write(0,2,'slug', bold)
    worksheet.write(0,3,'category_id', bold)

    #daya end

    workbook.close()
    output.seek(0)

    filename = 'django_sub_cat.xlsx'
    response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    response['Content-Disposition'] = 'attachment; filename=%s' % filename

    return response



@login_required(login_url='login')
@user_passes_test(check_role_vendor)
def generate_fourth_level_cat(request):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'remove_timezone': True})
    worksheet = workbook.add_worksheet('test')

    # cell_format = workbook.add_format({'bold': True, 'italic': True})


    # worksheet.write(0, 0, 'name',cell_format)
    # worksheet.write(0, 1, 'category',cell_format)

    #daya start

    bold = workbook.add_format({'bold': True})

    bold.set_pattern(1)  # This is optional when using a solid fill.
    bold.set_bg_color('orange')


    # Some data we want to write to the worksheet.
    reporte = Fourth_Level_Category.objects.all().order_by('id') #my model

    # Start from the first cell. Rows and columns are zero indexed.
    row = 1
    col = 0


    # Iterate over the data and write it out row by row.
    for linea in reporte:

        # naive_datetime = utc_to_local(linea.created_date)
        # naive_datetime = utc_datetime.replace(tzinfo=None)

        worksheet.write(row, col, linea.id)
        worksheet.write(row, col +  1, linea.name)
        worksheet.write(row, col +  2, linea.slug)

        worksheet.write(row, col +  3 ,linea.sub_category_id)
        row += 1
    
    worksheet.write(0,0,'id', bold)
    worksheet.write(0,1,'name', bold)
    worksheet.write(0,2,'slug', bold)
    worksheet.write(0,3,'category_id', bold)

    #daya end

    workbook.close()
    output.seek(0)

    filename = 'django_fourth_level_cat.xlsx'
    response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    response['Content-Disposition'] = 'attachment; filename=%s' % filename

    return response


@login_required(login_url='login')
@user_passes_test(check_role_vendor)               
def backup(request): 
    if request.method == 'POST' :
            file = request.FILES['excel_file']
            workbook = openpyxl.load_workbook(filename=file, read_only=True)
            # Get name of the first sheet and then open sheet by name
            first_sheet = workbook.get_sheet_names()[0]
            worksheet = workbook.get_sheet_by_name(first_sheet) 
            try:              
                for row in worksheet.iter_rows(): # Offset for header
                        row_dict = {  'category_id' : row[0],'product_name' : row[1],'featured_image': row[2],'image1' :row[3],
                                        'image2':row[4],'image3':row[5],'image4':row[6],'image5':row[7],'image6':row[8],'image7':row[9],'image8':row[10],
                                        'brand' : row[11],'product_information' : row[12],'variant' :row[13],'variant_title': row[14],'variant_color' : row[15],'variant_size' : row[16],
                                        'image_variant' : row[17],'variant_stock' : row[18],'variant_price' : row[19],'variant_discount' :row[20],'variant_max_allowed_quantity' :row[21],
                                        'stock' :row[22],'price' : row[23],'discount' :row[24],'weight' : row[25],'parcel_size_L' : row[26],'parcel_size_W' : row[27],
                                        'parcel_size_H' : row[28],'max_allowed_quantity' : row[29],'min_order_quantity' : row[30]
                                    }
                        
                        
                        if  Product.objects.get(product_name=row_dict['product_name'], brand=row_dict['brand']).count() < 2:
                
                            try:
                                    category_upload = Category_Upload.objects.get(id = row_dict['category_id'])

                                    if category_upload.main_category_id :
                                        main_category = Main_Category.objects.get(id=category_upload.main_category_id)
                                        
                                    if category_upload.category_id :
                                        category = Category.objects.get(id=category_upload.category_id)

                                    if category_upload.sub_category_id :
                                        sub_category = Sub_Category.objects.get(id=category_upload.sub_category_id)

                                    if category_upload.fourth_level_category_id :
                                        fourth_level_category = Fourth_Level_Category.objects.get(id=category_upload.fourth_level_category_id)
                            
                            except Category.DoesNotExist:
                                    pass

                            # Now, create the product
                            Product.objects.create(
                                main_category=main_category,
                                category=category,
                                sub_category=sub_category,
                                fourth_level_category=fourth_level_category,
                                product_name = row_dict['product_name'],
                                featured_image = row_dict['featured_image'],
                                brand = row_dict['brand'],
                                product_information = row_dict['product_information'],
                                variant = row_dict['variant'],
                                variant_stock = row_dict['stock'],
                                price = row_dict['price'],
                                discount = row_dict['discount'],
                                weight = row_dict['weight'],
                                parcel_size_L = row_dict['parcel_size_L'],
                                parcel_size_W = row_dict['parcel_size_W'],
                                parcel_size_H = row_dict['parcel_size_H'],
                                max_allowed_quantity = row_dict['max_allowed_quantity'],
                                min_order_quantity = row_dict['min_order_quantity']
                            )

                            product = Product.objects.get(product_name=row_dict['product_name'])

                            Variants.objects.create(
                                title = row_dict['variant_title'],
                                color = row_dict['variant_color'],
                                size = row_dict['variant_size'],
                                image_variant = row_dict['image_variant'],
                                variant_stock = row_dict['variant_stock'],
                                variant_price = row_dict['variant_price'],
                                variant_discount =row_dict['variant_discount'],
                                variant_max_allowed_quantity = row_dict['variant_max_allowed_quantity']
                            )

                        
                            Images.objects.create(product=product, image=row_dict['image1'])
                            Images.objects.create(product=product, image=row_dict['image2'])
                            Images.objects.create(product=product, image=row_dict['image3'])
                            Images.objects.create(product=product, image=row_dict['image4'])
                            Images.objects.create(product=product, image=row_dict['image5'])
                            Images.objects.create(product=product, image=row_dict['image6'])
                            Images.objects.create(product=product, image=row_dict['image7'])
                            Images.objects.create(product=product, image=row_dict['image8'])

                            upload_cnt+=1
                            
                            
                ProductRecord.objects.create(file_name = file,product=upload_cnt,status='Successful')
                messages.success(request,"Successful" ,extra_tags="saveexcel")
            except :
               messages.error(request,_('Error'),extra_tags="excelerror")

    return render(request, 'mass_upload/bulk_product.html')

# bulk_update
#           product.product_name = row_dict['product_name']
#                         product.featured_image = row_dict['featured_image']

#                         product.brand = row_dict['brand']
#                         product.product_information = row_dict['product_information']
#                         product.variant = row_dict['variant']

#                         product.stock = row_dict['stock'],
#                         product.price = row_dict['price'],
#                         product.discount = row_dict['discount'],
#                         product.weight = row_dict['weight'],
#                         product.parcel_size_L = row_dict['parcel_size_L'],
#                         product.parcel_size_W = row_dict['parcel_size_W'],
#                         product.parcel_size_H = row_dict['parcel_size_H'],
#                         product.max_allowed_quantity = row_dict['max_allowed_quantity'],
#                         product.min_order_quantity = row_dict['min_order_quantity']
#                         product.save()
            


def bulk_category(request):
     pass